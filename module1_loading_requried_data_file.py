from tkinter import filedialog
from openpyxl import load_workbook,Workbook

filename = filedialog.askopenfilename(defaultextension=".xlsx", filetypes=[("Excel File","*.xlsx")])
wb = load_workbook(filename=filename)
ws = wb.active

for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
    if row[0] is None:
        break
    print(row[0], row[1])