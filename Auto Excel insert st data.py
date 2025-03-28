# Automating excel for creating a file that contain students data that is extracted by the program
import os
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,PatternFill,Alignment

# from openpyxl.utils import get_column_letter
data = [{'Name': 'NITISH KUMAR', 'Registraion No': 2312051614, 'Roll No ': 6071580, 'Total Marks': '', 'Re appear ': 'Reappears : BCA201'}, {'Name': 'OMHARI MISHRA', 'Registraion No': 2312051533, 'Roll No ': 
6071581, 'Total Marks': '', 'Re appear ': 'Reappears : BCA201 BCA202 BCA203 BCA204 BCA205'}, {'Name': 'PANKAJ', 'Registraion No': 2312051459, 'Roll No ': 6071582, 'Total Marks': '', 'Re appear ': 'Reappears : BCA201 BCA202 BCA203 BCA204 BCA205'}, {'Name': 'PARAS KHANDELWAL', 'Registraion No': 2312051518, 'Roll No ': 6071583, 'Total Marks': '372', 'Re appear ': ''}, {'Name': 'PARIKSHIT', 'Registraion No': 2312051492, 'Roll No ': 6071584, 'Total Marks': '', 'Re appear ': 'Reappears : BCA201 BCA202 BCA203 BCA204 BCA205'}, {'Name': 'PAWAN KUMAR', 'Registraion No': 2312051656, 'Roll No ': 6071585, 'Total Marks': '', 'Re appear ': 'Reappears : BCA203'}, {'Name': 'PAWAN PANDEY', 'Registraion No': 2312051460, 'Roll No ': 6071586, 'Total Marks': '341', 'Re appear ': ''}, {'Name': 'PRABHAKAR PANDEY', 'Registraion No': 2312051462, 'Roll No ': 6071587, 'Total Marks': '272', 'Re appear ': ''}, {'Name': 'PRADEEP YADAV', 'Registraion No': 2312051464, 'Roll No ': 6071588, 'Total Marks': '', 'Re appear ': 'Reappears : BCA202 BCA204 BCA205'}, {'Name': 'PRANAY GADDI', 'Registraion No': 2312051525, 'Roll No ': 6071589, 'Total Marks': '', 'Re appear ': 'Reappears : BCA201 BCA202 BCA203 BCA204'}, {'Name': 'PRASHANT', 'Registraion No': 2312051782, 'Roll No ': 6071590, 'Total Marks': '', 'Re appear ': 'Reappears : BCA202'}, {'Name': 'PRASHANT KUMAR', 'Registraion No': 2312051463, 'Roll No ': 6071591, 'Total Marks': '266', 'Re appear ': ''}, {'Name': 'PREM CHAUHAN', 'Registraion No': 2312051465, 'Roll No ': 6071592, 'Total Marks': '251', 'Re appear ': ''}, {'Name': 'PRINCE', 'Registraion No': 2312051676, 'Roll No ': 6071593, 'Total Marks': '289', 'Re appear ': ''}, {'Name': 'PRINCE KUMAR', 'Registraion No': 2312051466, 'Roll No ': 6071594, 'Total Marks': '', 'Re appear ': 'Reappears : BCA205'}, {'Name': 'PRINCE PRAKASH', 'Registraion No': 2312051713, 'Roll No ': 6071595, 'Total Marks': '', 'Re appear ': 'Reappears : BCA201 BCA202 BCA203 BCA204'}, {'Name': 'PRIYAM TIWARI', 'Registraion No': 2312051717, 'Roll No ': 6071597, 'Total Marks': '', 'Re appear ': 'Reappears : BCA203'}, {'Name': 'PRIYANSHU', 'Registraion No': 2312051552, 'Roll No ': 6071598, 'Total Marks': '296', 'Re appear ': ''}, {'Name': 'PRIYANSHU KUMAR', 'Registraion No': 2312051467, 'Roll No ': 6071599, 'Total Marks': '', 'Re appear ': 'Reappears : BCA201 BCA202 BCA203 BCA204 BCA205'}, {'Name': 'PUNEET DAHIYA', 'Registraion No': 2312051530, 'Roll No ': 6071600, 'Total Marks': '', 'Re appear ': 'Reappears : BCA203'}, {'Name': 'RAHUL', 'Registraion No': 2312051654, 'Roll No ': 6071601, 'Total Marks': '', 'Re appear ': 'Reappears : BCA202 BCA203 BCA204'}, {'Name': 'RAHUL KUMAR SINGH', 'Registraion No': 2312051502, 'Roll No ': 6071602, 'Total Marks': '', 'Re appear ': 'Reappears : BCA202'}, {'Name': 'RAHUL SINGH', 'Registraion No': 2312051468, 'Roll No ': 6071603, 'Total Marks': '329', 'Re appear ': ''}, {'Name': 'RAJA KUMAR', 'Registraion No': 2312051469, 'Roll No ': 6071604, 'Total Marks': '', 'Re appear ': 'Reappears : BCA201 BCA202 BCA203 BCA204'}, {'Name': 'RAJA KUMAR SINGH', 'Registraion No': 2312051616, 'Roll No ': 6071605, 'Total Marks': '', 'Re appear ': 'Reappears : BCA202 BCA204'}, {'Name': 'RAJNEESH GUPTA', 'Registraion No': 2312051779, 'Roll No ': 6071606, 'Total Marks': '', 'Re appear ': 'Reappears : BCA201 BCA202 BCA203 BCA204 BCA205'}, {'Name': 'RAJU', 'Registraion No': 2312051470, 'Roll No ': 6071607, 'Total Marks': '391', 'Re appear ': ''}, {'Name': 'RAM BABU SINGH', 'Registraion No': 2312051510, 'Roll No ': 6071608, 'Total Marks': '', 'Re appear ': 'Reappears : BCA201 BCA202 BCA203 BCA204'}, {'Name': 'RANJEET KUMAR', 'Registraion No': 2312051688, 'Roll No ': 6071609, 'Total Marks': '', 'Re appear ': 'Reappears : BCA201 BCA202 BCA203 BCA204 BCA205'}, {'Name': 'RISHAB HARSANA', 'Registraion No': 2312051617, 'Roll No ': 6071610, 'Total Marks': '', 'Re appear ': 'Reappears : BCA201 BCA202 BCA203 BCA204'}, {'Name': 'RITIK YADAV', 'Registraion No': 2312051471, 'Roll No ': 6071611, 'Total Marks': '333', 'Re appear ': ''}, {'Name': 'ROBIN', 'Registraion No': 2312051618, 'Roll No ': 6071612, 'Total Marks': '283', 'Re appear ': ''}, {'Name': 'ROHIT', 'Registraion No': 2312051546, 'Roll No ': 6071613, 'Total Marks': '', 'Re appear ': 'Reappears : BCA201 BCA202 BCA203 BCA204'}, {'Name': 'ROHIT KUMAR PRAJAPATI', 'Registraion No': 2312051776, 'Roll No ': 6071614, 'Total Marks': '', 'Re appear ': 'Reappears : BCA201 BCA202 BCA203 BCA204 BCA205'}, {'Name': 'ROHIT KUMAR UJJWAL', 'Registraion No': 2312051526, 'Roll No ': 6071615, 'Total Marks': '325', 'Re appear ': ''}, {'Name': 'RUDRANSHU PANDEY', 'Registraion No': 2312051599, 'Roll No ': 6071616, 'Total Marks': '', 'Re appear ': 'Reappears : BCA202 BCA203 BCA204'}, {'Name': 'SACHIN', 'Registraion No': 2312051805, 'Roll No ': 6071617, 'Total Marks': '', 'Re appear ': 'Reappears : BCA201 BCA202 BCA203 BCA204 BCA205'}, {'Name': 'SACHIN YADAV', 'Registraion No': 2312051621, 'Roll No ': 6071619, 'Total Marks': '', 'Re appear ': 'Reappears : BCA201 BCA202 BCA203 BCA204 BCA205'}]

# Create a new workbook
file_name = "students result data .xlsx"
if os.path.exists(file_name):
    # loding existing file
    wb = load_workbook(file_name)
else:

    wb = Workbook()

if "sheet1" in wb.sheetnames:
    ws = wb["sheet1"]
else:
    ws = wb.active
    ws.title = "sheet1"

# Merging Cells for Header of the File
ws['A1'].fill = PatternFill(start_color='c6efce',end_color='c6efce',fill_type='solid') # changing cell bg color to light green
ws['A1'] = "BCA-III Sem Result"
ws['A1'].font = Font(name="Calibri",size=15,bold=True,color="006100")
ws['A1'].alignment = Alignment(horizontal='center',vertical='center')

# Writing the Header of the file like name Regis, roll , etc
headers = ["Sr.No","Name","Registration No","Roll No","Total Marks","Reappear"]
cols = 1
for head in headers:
    cols_char = get_column_letter(cols)
    ws[cols_char+str(2)] = head
    cols+=1


#Entering data logic for  Inserting data in excel

rows_counter = 3
column_counter = 2
for dict in data:
    for di in dict:
        cols_char = get_column_letter(column_counter)
        ws[cols_char+str(rows_counter)] = dict[di]
        if column_counter == 6:
            column_counter = 1 
        column_counter += 1
    rows_counter += 1




wb.save("students result data .xlsx")


