# Required framework and library
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,PatternFill,Alignment
import time,os #Inbuilt Modules
from tkinter import messagebox as msg
#Automate the task using selenium"
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


#initializing a file which track the of the number of row for automate excel data
current_row = 4
tracker_file = "row_status.txt"
def track_row():
    if os.path.exists(tracker_file):
        with open(tracker_file,'r') as file:
            row = file.readline().strip()
    else:
        with open(tracker_file,'w') as file:
            file.write(str(current_row))
        row = str(current_row)
    return row

#initializing Excel File Here
excel_file = "student_result_data.xlsx"
if os.path.exists(excel_file):
    wb = load_workbook(excel_file)
else:
    wb = Workbook()

if "sheet1" in wb.sheetnames:
    ws = wb['sheet1']
else:
    ws = wb.active
    ws.title = "sheet1"

# --- Title Header (Row 1) ---
ws.merge_cells('A1:N1')
ws['A1'].fill = PatternFill(start_color="8DFAB1",end_color="8DFAB1",fill_type="solid")
ws['A1'] = "BCA Result 2025 Sem - IV"
ws['A1'].font = Font("Times New Roman",size="16",bold=True,color="006100")
ws['A1'].alignment = Alignment(horizontal="center",vertical="center")

# --- Section Headers (Row 2) ---
section_headers = [
    {"cell": "A2:E2", "text": "Students Details"},
    {"cell": "G2:K2", "text": "Marks in Each Subjects"},
    {"cell": "M2:N2", "text": "Result"}]

for section in section_headers:
    ws.merge_cells(section["cell"])
    cell_ref = section["cell"].split(":")[0]  # Get the top-left cell of the merged area
    ws[cell_ref] = section["text"]
    ws[cell_ref].fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    ws[cell_ref].font = Font(name="Century", size=14, color="006100")
    ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center")


# Adding Header of the Columns Formatting
col_head = ["Sr.No","Name","Father Name","Registration No","Roll No","","BCA206","BCA207","BCA208","BCA209","BCA210","","Total Marks","Result"]

cols = 1
for header in col_head:
    cols_char = get_column_letter(cols)
    ws[cols_char+str(3)] = header
    cols += 1


# Set up the WebDriver (Edge in this example)
options = Options()
options.add_experimental_option('excludeSwitches',['enable-logging'])  # Suppress DevTools message
driver = webdriver.Edge(options=options)
wait = WebDriverWait(driver,10)
# Navigate to the result portal
driver.get("https://result.mdu.ac.in/postexam/result.aspx")


#"Enter Your data Here 👇👇👇"
      #Data format (registration,Roll No)
student_details = [(2312051470,6071607)]
process_data = []
file_name = "Processed_entries.txt"
#Updating Current row
current_row = int(track_row())

try:
    if os.path.exists(file_name):
        with open(file_name,"r") as file:
            process_data = list(map(int,file.read().split()))

except Exception as e:
    msg.ERROR(e)

finally:
    with open(file_name,"a") as file:
        for Registration_No,Roll_no in student_details:
            try:
                if Registration_No in process_data:
                    continue
                else:
                    
                    # Process of Extracting Result
                    driver.find_element(By.ID, "txtRegistrationNo").send_keys(Registration_No)  # Registration Number one by one
                    driver.find_element(By.ID, "txtRollNo").send_keys(Roll_no)  # Roll Number one by one
                    # Submit the form
                    wait.until(EC.element_to_be_clickable((By.ID,"cmdbtnProceed"))).click()
                    wait.until(EC.element_to_be_clickable((By.ID,"imgComfirm"))).click()
                    wait.until(EC.element_to_be_clickable((By.ID,"rptMain_ctl01_lnkView"))).click()
                    wait.until(EC.visibility_of_element_located((By.ID,"lblStudentName")))
                    student_name = driver.find_element(By.ID,"lblStudentName").text
                    father_name = driver.find_element(By.ID,"lblFatherName").text

                    #Extracting Subjects Marks
                    subject_marks = [driver.find_element(By.XPATH, f"//table//tr[{i}]//td[8]").text for i in range(2, 7)]
                   #random_data = driver.find_element(By.XPATH("//table//tr[2]//td[8]"))#just for example a simple way for extracting random data
                    bca201, bca202, bca203, bca204, bca205 = subject_marks
                    
                    #Result
                    total_marks = driver.find_element(By.ID,"rptMarks_ctl06_lblTotal").text
                    result = driver.find_element(By.ID,"lblresult").text
            
            # Storing data in Excel for each row
                    data = [current_row-3,student_name,father_name,Registration_No,Roll_no,"",bca201,bca202,bca203,bca204,bca205,"",total_marks,result]
                    for char in range(1,15):
                        char_num = get_column_letter(char)
                        ws[char_num+str(current_row)] = data[char-1]
                        with open(tracker_file,"w") as file2:
                            file2.write(str(current_row))
                    current_row += 1
                    
                    #Just for Checking Current status
                    print(f"{student_name = }")
                    #Updating registraion_no in processed_entries file
                    file.write(str(Registration_No) + "\n")
                    driver.refresh() #Reloading website for repreating process

            except Exception as e:
                wb.save(excel_file)
                print(f"Error is occured In {Registration_No} and \nUnexpected Error: Please ensure good internet connect or try again")
                msg.showerror("process Interupted",f"{e}")
                exit()
                    
        wb.save(excel_file)
#updating the last successful record row
    with open(tracker_file,'w') as track_update:
        track_update.write(str(current_row))
    # Display info when process complete
    msg.showinfo("Success","All the data fatched successfully")
    # Close the browser
    driver.quit()
